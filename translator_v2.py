from tkinter import Tk, Frame, Button, filedialog, StringVar, OptionMenu, Label, Text, Radiobutton, messagebox
from openpyxl import load_workbook, Workbook
import deepl


class Translator:

	SOURCE_LANG_LIST = ['Autodetect', 'CS', 'DA', 'DE', 'EL', 'EN', 'ES', 'ET', 'FI', 'FR', 'HU', 'IT', 'JA', 'LT', 'LV',
						'NL', 'PL', 'PT', 'RO', 'RU', 'SK', 'SL', 'SV', 'ZH']
	TARGET_LANG_LIST = ['BG', 'CS', 'DA', 'DE', 'EL', 'EN-GB', 'EN-US', 'ES', 'ET', 'FI', 'FR', 'HU', 'IT', 'JA',
						'LT', 'LV', 'NL', 'PL', 'PT-PT', 'PT-BR', 'RO', 'RU']
	AUTHENTICATION_KEY = ''

	def __init__(self):
		self._window = Tk()
		self._window.resizable(False, False)
		self._window.geometry('380x270')
		self._window.eval("tk::PlaceWindow . center")
		self._window.title("Deepl API Translator")
		self.build_frames()
		self.build_radio_buttons()
		self.build_buttons()
		self.build_labels()
		self.build_menu()
		self.build_info()
		self.translator = deepl.Translator(auth_key=Translator.AUTHENTICATION_KEY)

	def create_frame(self, master, row, column, rowspan, columnspan):
		frame = Frame(master=master, padx=2, pady=2)
		frame.grid(row=row, column=column, columnspan=columnspan, rowspan=rowspan, sticky="nwse")
		return frame

	def create_buttons(self, frame, text, row, column, command):
		button = Button(master=frame, text=text, command=command)
		button.grid(row=row, column=column)
		return button

	def create_label(self, master, row, column, rowspan, columnspan, wraplength=None, justify=None):
		self.file_name = StringVar()
		self.file_name.set('Nevybrán žádný soubor')
		label = Label(master=master, text=self.file_name.get())
		label.config(wraplength=wraplength, justify=justify)
		label.grid(row=row, column=column, rowspan=rowspan, columnspan=columnspan)
		return label

	def create_menu(self, master, default_value, langs, row, column, rowspan, columnspan, lang):
		self.lang_variable = StringVar(self._window)
		self.lang_variable.set(default_value) # default value
		menu = OptionMenu(master, self.lang_variable, *langs, command=lang)
		menu.config(width=12)
		menu.grid(row=row, column=column, rowspan=rowspan, columnspan=columnspan)
		return menu

	def build_frames(self):
		self.main_frame = self.create_frame(self._window, 0, 0, 6, 4).config(width=380, height=240)
		self.topleft_frame = self.create_frame(self.main_frame, 0, 0, 2, 1)
		self.topright_frame = self.create_frame(self.main_frame, 0, 1, 1, 1)
		self.right_frame = self.create_frame(self.main_frame, 1, 1, 5, 1)
		self.middleleft_frame = self.create_frame(self.main_frame, 1, 0, 2, 1)
		self.bottomleft_frame = self.create_frame(self.main_frame, 3, 0, 3, 1)

	def build_labels(self):
		self.selected_file_label = self.create_label(self.topright_frame, 1, 0, 1, 1, 250, 'left')

	def build_menu(self):
		self.source_lang_menu = self.create_menu(self.bottomleft_frame,'Zdrojový jazyk',
												 Translator.SOURCE_LANG_LIST, 1, 0, 1, 1, self.source_lang_selection)
		self.target_lang_menu = self.create_menu(self.bottomleft_frame,'Cílový jazyk',
												 Translator.TARGET_LANG_LIST, 2, 0, 1, 1, self.target_lang_selection)

	def build_buttons(self):
		self.load_file_button = self.create_buttons(self.topleft_frame, 'Načíst soubor', 0, 0, self.load_file)
		self.translate_button = self.create_buttons(self.bottomleft_frame, 'Přeložit', 3, 0, self.translate)
		self.check_limit_button = self.create_buttons(self.bottomleft_frame, 'Zkontroluj limit', 4, 0, self.check_limit)

	def build_radio_buttons(self):
		buttons = [('Pouze popisy', '1'), ('Popisy a nazvy', '2')]
		self.r_button_variable = StringVar()
		self.r_button_variable.set('1')
		for text, mode in buttons:
			b = Radiobutton(self.middleleft_frame, text=text, variable=self.r_button_variable, value=mode)
			b.grid()
		return buttons

	def build_info(self):
		self.info_box = Text(self.right_frame)
		self.info_box.config(height=14, width=40, font='Arial 8')
		self.info_box.grid()
		return self.info_box

	def load_file(self):
		file = filedialog.askopenfilename(filetypes=[('Excel soubory', '*.xlsx')])
		user_selected_file = load_workbook(file, data_only=True)
		self.file_name.set(file)
		self.selected_file_label.config(text=self.file_name.get())
		self.ws1 = user_selected_file.active
		return self.ws1

	def source_lang_selection(self, source_lang_selection):
		if source_lang_selection == 'Autodetect':
			self.source_lang = None
			return self.source_lang
		else:
			self.source_lang = source_lang_selection
			return self.source_lang

	def target_lang_selection(self, target_lang_selection):
		self.target_lang = target_lang_selection
		return self.target_lang

	def check_limit(self):
		usage = str(self.translator.get_usage())
		values = [int(s) for s in usage.split() if s.isdigit()]
		percentage_usage = round((values[0] / values[1]) * 100, 2)
		self.info_box.delete('0.0', 'end-1c')
		self.info_box.config(background='white')
		self.info_box.insert(1.0, f'Aktuálně využito {values[0]} ({percentage_usage}%) z celkového počtu {values[1]} znaků.\n')
		self.info_box.insert(2.0, f'Ještě možno využít {values[1] - values[0]} znaků.')
		self.info_box.update()

	def translate(self):
		exceptions = []
		results = {}
		self.info_box.config(background='white')
		self.info_box.delete("0.0", "end-1c")
		if self.r_button_variable.get() == '1':
			for row in self.ws1.iter_rows(max_col=2, min_row=2, values_only=True):
				self.info_box.insert(1.0, f'Překládám {row[0]}\n')
				self.info_box.update()
				try:
					results[row[0]] = self.translator.translate_text(row[1], source_lang=self.source_lang,
																	 target_lang=self.target_lang).text
				except deepl.exceptions.QuotaExceededException:
					exceptions.append('quota')
					messagebox.showinfo('Informace', 'Quota exceeded')
					break
				except deepl.exceptions.TooManyRequestsException:
					exceptions.append('requests')
					messagebox.showinfo('Informace', 'Too many requests. Try again later.')
					break

			final_excel = Workbook()
			fe_ws1 = final_excel.active

			for values in results.items():
				fe_ws1.append([str(values[0]), str(values[1])])

			if 'requests' in exceptions:
				final_excel.save(f'preklad_{self.target_lang}.xlsx')
				self.info_box.config(background='orange')
				self.info_box.insert(1.0, 'Překlad dokončen jen z části\n\n')
				self.info_box.update()
			elif 'quota' in exceptions:
				final_excel.save(f'preklad_{self.target_lang}.xlsx')
				self.info_box.config(background='orange')
				self.info_box.insert(1.0, 'Překlad dokončen jen z části\n\n')
				self.info_box.update()
			else:
				final_excel.save(f'preklad_{self.target_lang}.xlsx')
				self.info_box.config(background='green')
				self.info_box.insert(1.0, 'Překlad dokončen\n\n')
		elif self.r_button_variable.get() == '2':
			for row in self.ws1.iter_rows(max_col=3, min_row=2, values_only=True):
				temp = []
				self.info_box.insert(1.0, f'Překládám {row[0]}\n')
				self.info_box.update()
				try:
					temp.append(self.translator.translate_text(row[1], source_lang=self.source_lang,
															   target_lang=self.target_lang).text)
					temp.append(self.translator.translate_text(row[2], source_lang=self.source_lang,
															   target_lang=self.target_lang).text)
					results[row[0]] = temp
				except deepl.exceptions.QuotaExceededException:
					exceptions.append('quota')
					messagebox.showinfo('Informace', 'Quota exceeded')
					break
				except deepl.exceptions.TooManyRequestsException:
					exceptions.append('requests')
					messagebox.showinfo('Informace', 'Too many requests. Try again later.')
					break

			final_excel = Workbook()
			fe_ws1 = final_excel.active
			for values in results.items():
				fe_ws1.append([str(values[0]), str(values[1][0]), str(values[1][1])])

			if 'requests' in exceptions:
				final_excel.save(f'preklad_{self.target_lang}.xlsx')
				self.info_box.config(background='orange')
				self.info_box.insert(1.0, 'Překlad dokončen jen z části\n\n')
				self.info_box.update()
			elif 'quota' in exceptions:
				final_excel.save(f'preklad_{self.target_lang}.xlsx')
				self.info_box.config(background='orange')
				self.info_box.insert(1.0, 'Překlad dokončen jen z části\n\n')
				self.info_box.update()
			else:
				final_excel.save(f'preklad_{self.target_lang}.xlsx')
				self.info_box.config(background='green')
				self.info_box.insert(1.0, 'Překlad dokončen\n\n')

	def start_window(self):
		self._window.mainloop()


if __name__ == "__main__":
	tr = Translator()
	tr.start_window()
